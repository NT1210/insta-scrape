from scrape import scrape
from time import sleep
from openpyxl import load_workbook

def loadExcel(date_today_formatted):
    wb = load_workbook(filename = './template.xlsx')
    ws = None

    try:
        ws = wb[date_today_formatted]
    except KeyError:
        print("Sheet does not exist. Please add a new sheet corresponding to the designated month")
        return
    
    return wb, ws

def processExcel(date_for_excel, wb, ws, progress):
    target_column = 0
    for i in range(5, 36):
        if str(ws.cell(row=2, column=i).value).split(" ")[0] == date_for_excel:
            target_column = i
            break
        else:
            continue

    progress_count = 0
    last_row = ws.max_row
    for i in range(3, last_row+1):
        target_id = ws.cell(row=i, column=3).value
        scraped_data = {}

        if target_id != None:
            scraped_data = scrape(target_id)

            temp_obj = {
                "name": scraped_data["username"],
                "posts": scraped_data["edge_owner_to_timeline_media"]["count"],
                "follower": scraped_data["edge_followed_by"]["count"],
                "following": scraped_data["edge_follow"]["count"]
            }
            
            for j in range(3):
                match j:
                    case 0:
                        ws.cell(row=i+j, column=target_column).value = temp_obj["posts"]
                    case 1:
                        ws.cell(row=i+j, column=target_column).value = temp_obj["follower"]
                    case 2:
                        ws.cell(row=i+j, column=target_column).value = temp_obj["following"]
            
            progress.update(progress_count)
            sleep(3)
            progress_count = progress_count+1
        else:
            continue

        
    
    file_for_output = "./" + date_for_excel + "" + ".xlsx" #./2024-07-01.xlsx
    wb.save(file_for_output)

    return progress